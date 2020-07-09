import openpyxl as xl
import random
wb = xl.load_workbook('ISR.xlsx')
sheet = wb['Sheet1']


def Mod(row, colum, value):
    location = sheet.cell(row, colum)
    location.value = value
    location = sheet.cell(row, colum)
    wb.save("ISR.xlsx")
y=1
for v in range (1,40):
    z = random.randint(87, 100)

    m = str(v + 1)
    print("TRAIL" + m)
    wb.create_sheet("TRAIL" + str(m) )
    Mod(1, 1,  ("Tile Number"))
    Mod(1, 2,  ("% of accurate sound picked up"))
    Mod(1, 3,  ("Meters away from Mic"))

    c = 1


    for x in range (2,z):
        Mod(x, 1 ,(x-1))
        if c == z:
            Mod(x, 2, ("0%"))
        if c < z :
            Mod(x,2,("100%"))
        Mod(x,3,5 * (x-1))

        c = c+1


        sheet = wb['TRAIL' + (m)]


wb.save("ISR.xlsx")
